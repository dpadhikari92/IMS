import datetime
import decimal
from random import choice
from django.contrib.auth.decorators import login_required, permission_required
from django import apps
from django.db.models import Q
from django.db.models import Sum
from django.forms import formset_factory, modelformset_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import (
    View, 
    ListView,
    CreateView,
    UpdateView,
    DeleteView
)
from django.http import FileResponse, HttpResponse
import io
from reportlab.lib.pagesizes import letter,landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from io import BytesIO
from reportlab.pdfgen import canvas
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import (
    FGSFG,
    PurchaseBill, 
    Supplier, 
    PurchaseItem,
    PurchaseBillDetails,
    SaleBill,  
    SaleItem,
    SaleBillDetails,
    BOM,
    BOMRawMaterial,
    Production,   
    FGSFG,
    ProductionFG,
    Leadtimesfg,
    Leadtimefg,
    SfgBom,
    Sfgfinal,
    sfgproduction,
    FGSFGNEW,
    fgproduction,
    RawMaterialEntry,
    MultipleSFG,
)

from .forms import (
    PurchaseItemForm,
    SelectSupplierForm, 
    PurchaseItemFormset,
    PurchaseDetailsForm, 
    SupplierForm, 
    SaleForm,
    SaleItemFormset,
    SaleDetailsForm, 
    
)
from inventory.models import Stock
from . import models

from django_filters.views import FilterView
from .filters import BOMFilter
from django.db import transaction  






# shows a lists of all suppliers
class SupplierListView(ListView):
    model = Supplier
    template_name = "suppliers/suppliers_list.html"
    queryset = Supplier.objects.filter(is_deleted=False)
    paginate_by = 10



# used to add a new supplier
class SupplierCreateView(SuccessMessageMixin, CreateView):
    model = Supplier
    form_class = SupplierForm
    success_url = '/transactions/suppliers'
    success_message = "Supplier has been created successfully"
    template_name = "suppliers/edit_supplier.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = 'New Supplier'
        context["savebtn"] = 'Add Supplier'
        return context     


# used to update a supplier's info
class SupplierUpdateView(SuccessMessageMixin, UpdateView):
    model = Supplier
    form_class = SupplierForm
    success_url = '/transactions/suppliers'
    success_message = "Supplier details has been updated successfully"
    template_name = "suppliers/edit_supplier.html"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = 'Edit Supplier'
        context["savebtn"] = 'Save Changes'
        context["delbtn"] = 'Delete Supplier'
        return context


# used to delete a supplier
class SupplierDeleteView(View):
    template_name = "suppliers/delete_supplier.html"
    success_message = "Supplier has been deleted successfully"

    def get(self, request, pk):
        supplier = get_object_or_404(Supplier, pk=pk)
        return render(request, self.template_name, {'object' : supplier})

    def post(self, request, pk):  
        supplier = get_object_or_404(Supplier, pk=pk)
        supplier.is_deleted = True
        supplier.save()                                               
        messages.success(request, self.success_message)
        return redirect('suppliers-list')


# used to view a supplier's profile
class SupplierView(View):
    def get(self, request, name):
        supplierobj = get_object_or_404(Supplier, name=name)
        bill_list = PurchaseBill.objects.filter(supplier=supplierobj)
        page = request.GET.get('page', 1)
        paginator = Paginator(bill_list, 10)
        try:
            bills = paginator.page(page)
        except PageNotAnInteger:
            bills = paginator.page(1)
        except EmptyPage:
            bills = paginator.page(paginator.num_pages)
        context = {
            'supplier'  : supplierobj,
            'bills'     : bills
        }
        return render(request, 'suppliers/supplier.html', context)



class PurchaseView(ListView):
    model = PurchaseBill
    template_name = "purchases/purchases_list.html"
    context_object_name = 'bills'
    ordering = ['-billno']
    paginate_by = 4  # Number of records per page

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(purchase_item__purchase_code__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')

        # Get the current page number from the request's GET parameters
        page_number = self.request.GET.get('page')

        # Get the latest 10 records if the page number is not specified
        if not page_number:
            context['bills'] = context['bills'][:4]

        return context
    


# used to select the supplier
class SelectSupplierView(View):
    form_class = SelectSupplierForm
    template_name = 'purchases/select_supplier.html'

    def get(self, request, *args, **kwargs):                                    # loads the form page
        form = self.form_class
        return render(request, self.template_name, {'form': form})

    def post(self, request, *args, **kwargs):                                   # gets selected supplier and redirects to 'PurchaseCreateView' class
        form = self.form_class(request.POST)
        if form.is_valid():
            supplierid = request.POST.get("supplier")
            supplier = get_object_or_404(Supplier, id=supplierid)
            return redirect('new-purchase', supplier.pk)
        return render(request, self.template_name, {'form': form})


class PurchaseCreateView(View):
    template_name = 'purchases/new_purchase.html'

    def get(self, request, pk):
        formset = PurchaseItemFormset(request.GET or None)
        supplierobj = get_object_or_404(Supplier, pk=pk)
        context = {
            'formset': formset,
            'supplier': supplierobj,
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        formset = PurchaseItemFormset(request.POST)
        supplierobj = get_object_or_404(Supplier, pk=pk)

        if formset.is_valid():
            billobj = PurchaseBill(supplier=supplierobj)
            billobj.save()
            billdetailsobj = PurchaseBillDetails(billno=billobj)
            billdetailsobj.save()
            for form in formset:
                billitem = form.save(commit=False)
                billitem.billno = billobj
                stock = get_object_or_404(Stock, name=billitem.stock.name)
                billitem.totalprice = billitem.perprice * billitem.quantity
                stock.quantity += billitem.quantity
                stock.save()
                
               
                 
                billitem.save()
                stock.purchase_items.add(billitem)
            messages.success(request, "Purchased items have been registered successfully")
            return redirect('purchase-bill', billno=billobj.billno)
        formset = PurchaseItemFormset(request.GET or None)
        context = {
            'formset': formset,
            'supplier': supplierobj
        }
        return render(request, self.template_name, context)


class PurchaseUpdateView(View):
    
        template_name = 'purchases/update_purchase.html'  # Update with your template name

        
        def get(self, request, billno):
            billobj = get_object_or_404(PurchaseBill, billno=billno)
            PurchaseItemFormset = modelformset_factory(PurchaseItem, form=PurchaseItemForm, extra=0)
            formset = PurchaseItemFormset(queryset=billobj.purchase_item.all())
            context = {
                'formset': formset,
                'billno': billno,
            }
            return render(request, self.template_name, context)

        def post(self, request, billno):
            billobj = get_object_or_404(PurchaseBill, billno=billno)
            PurchaseItemFormset = modelformset_factory(PurchaseItem, form=PurchaseItemForm, extra=0)
            formset = PurchaseItemFormset(request.POST, queryset=billobj.purchase_item.all())

            if formset.is_valid():
                # Update existing items and quantities in the stock
                for form, old_item in zip(formset, billobj.purchase_item.all()):
                    new_item = form.save(commit=False)
                    old_quantity = old_item.quantity
                    old_stock = get_object_or_404(Stock, name=old_item.stock.name)
                    old_stock.quantity -= old_quantity
                    old_stock.save()

                    new_quantity = new_item.quantity
                    new_stock = get_object_or_404(Stock, name=new_item.stock.name)
                    new_stock.quantity += new_quantity
                    new_stock.save()

                    new_item.totalprice = new_item.perprice * new_quantity
                    new_item.save()

                    old_item.delete()

                # Add new items to the bill
                for form in formset:
                    new_item = form.save(commit=False)
                    new_quantity = new_item.quantity
                    new_stock = get_object_or_404(Stock, name=new_item.stock.name)
                    new_stock.quantity += new_quantity
                    new_stock.save()

                    new_item.totalprice = new_item.perprice * new_quantity
                    new_item.billno = billobj
                    new_item.save()

                messages.success(request, "Purchase items have been updated successfully")
                return redirect('purchase-bill', billno=billno)

            context = {
                'formset': formset,
                'billno': billno,
            }
            return render(request, self.template_name, context)


# used to download purchase report
from datetime import datetime
from django.template.loader import render_to_string
from openpyxl import Workbook


def purchase_report(request):
    if request.method == 'POST':
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')

        # Convert the date strings to datetime objects
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()

        # Fetch PurchaseBill instances within the date range
        purchase_bills = PurchaseBill.objects.filter(time__range=(from_date, to_date))

        # Fetch related PurchaseItem instances using the retrieved PurchaseBill instances
        purchase_items = PurchaseItem.objects.filter(billno__in=purchase_bills)

        # Calculate total quantity for each item
        item_totals = purchase_items.values('stock__name').annotate(total_quantity=Sum('quantity'))

        context = {'item_totals': item_totals, 'from_date': from_date, 'to_date': to_date}

        return render(request, 'purchases/purchase_report.html', context)

    return render(request, 'purchases/purchase_report.html')
    
    
    
# used to download purchase report
def download_purchase_report(request):
    if request.method == 'POST':
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')

        # Convert the date strings to datetime objects
        from_date = datetime.strptime(from_date, '%b. %d, %Y').date()
        to_date = datetime.strptime(to_date, '%b. %d, %Y').date()

        # Fetch PurchaseBill instances within the date range
        purchase_bills = PurchaseBill.objects.filter(time__range=(from_date, to_date))

        # Fetch related PurchaseItem instances using the retrieved PurchaseBill instances
        purchase_items = PurchaseItem.objects.filter(billno__in=purchase_bills)

        # Calculate total quantity for each item
        item_totals = purchase_items.values('stock__name').annotate(total_quantity=Sum('quantity'))

        # Generate and download the Excel report
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="purchase_report.xlsx"'

        wb = Workbook()
        ws = wb.active

        # Add From and To date
        ws['A1'] = f'From: {from_date}'
        ws['A2'] = f'To: {to_date}'

        # Add headers
        ws['A4'] = 'Item'
        ws['B4'] = 'Total Quantity (Kg)'

        # Add data
        row = 5
        for item_total in item_totals:
            ws.cell(row=row, column=1, value=item_total['stock__name'])
            ws.cell(row=row, column=2, value=item_total['total_quantity'])
            row += 1

        wb.save(response)
        return response



import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse
from datetime import datetime
from django.db.models import Sum, Case, When, Value, F, FloatField
from django.db.models.functions import Coalesce, Cast
from .models import Stock
from django.shortcuts import render

def inventory_report(request):
    date_str = request.GET.get('date')
    inventory_items = []

    if date_str:
        date = datetime.strptime(date_str, '%Y-%m-%d')
        inventory_query = Stock.objects.filter(is_deleted=False)
        inventory_items = inventory_query.annotate(
            total_quantity=Coalesce(Sum(
                Case(
                    When(purchase_items__billno__purchasedetailsbillno__invoice__lte=date, then=F('purchase_items__quantity')),
                    default=Value(0),
                    output_field=FloatField()
                )
            ), Value(0), output_field=FloatField())
        )

        for item in inventory_items:
            # Deduct production quantities based on the production date and BOM relationships
            production_qty = Production.objects.filter(bom__raw_materials=item, production_date__lte=date).aggregate(
                total_production=Coalesce(Sum('quantity'), Value(0), output_field=FloatField())
            )['total_production']
            
            # Deduct SFG (semi-finished goods) production quantities based on the production date and BOM relationships
            sfg_production_qty = fgproduction.objects.filter(bom__sfg__raw_materials=item, production_date__lte=date).aggregate(
                total_production=Coalesce(Sum('quantity'), Value(0), output_field=FloatField())
            )['total_production']
            
            item.total_quantity -= production_qty + sfg_production_qty
    
    context = {
        'date': date_str,
        'inventory_items': inventory_items,
    }
    return render(request, 'purchases/inventory_report.html', context)

    
def download_excel(request):
    date_str = request.GET.get('date')
    

    inventory_items = []
    if date_str:
        date = datetime.strptime(date_str, '%Y-%m-%d')
        inventory_query = Stock.objects.filter(is_deleted=False)
        inventory_items = inventory_query.annotate(
            total_quantity=Coalesce(Sum(
                Case(
                    When(purchase_items__billno__purchasedetailsbillno__invoice__lte=date, then=F('purchase_items__quantity')),
                    default=Value(0),
                    output_field=FloatField()
                )
            ), Value(0), output_field=FloatField())
        )

        for item in inventory_items:
            # Deduct production quantities based on the production date and BOM relationships
            production_qty = Production.objects.filter(bom__raw_materials=item, production_date__lte=date).aggregate(
                total_production=Coalesce(Sum('quantity'), Value(0), output_field=FloatField())
            )['total_production']
            
            # Deduct SFG (semi-finished goods) production quantities based on the production date and BOM relationships
            sfg_production_qty = fgproduction.objects.filter(bom__sfg__raw_materials=item, production_date__lte=date).aggregate(
                total_production=Coalesce(Sum('quantity'), Value(0), output_field=FloatField())
            )['total_production']
            
            item.total_quantity -= production_qty + sfg_production_qty
    


    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write column headers
    column_headers = ['Item Name', 'inventory level',]
    for col_num, column_title in enumerate(column_headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f'{col_letter}1']
        cell.value = column_title
        cell.font = Font(bold=True)

    # Populate data rows
    for row_num, item in enumerate(inventory_items, start=2):
        worksheet[f'A{row_num}'] = item.name

        # Format total_quantity and total_quantity_date2 with trailing zeroes
        decimal_places = 4
        total_quantity = decimal.Decimal(item.total_quantity).quantize(decimal.Decimal(f'0.{decimal_places * "0"}'))
        

        worksheet[f'B{row_num}'] = total_quantity
      

    # Create an HTTP response with the Excel file content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=inventory_report.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response


# used to delete a bill object
class PurchaseDeleteView(SuccessMessageMixin, DeleteView):
    model = PurchaseBill
    template_name = "purchases/delete_purchase.html"
    success_url = '/transactions/purchases'
    
    def delete(self, *args, **kwargs):
        self.object = self.get_object()
        items = PurchaseItem.objects.filter(billno=self.object.billno)
        for item in items:
            stock = get_object_or_404(Stock, name=item.stock.name)
            if stock.is_deleted == False:
                stock.quantity -= item.quantity
                stock.save()
        messages.success(self.request, "Purchase bill has been deleted successfully")
        return super(PurchaseDeleteView, self).delete(*args, **kwargs)




# shows the list of bills of all sales 
class SaleView(ListView):
    model = SaleBill
    template_name = "sales/sales_list.html"
    context_object_name = 'bills'
    ordering = ['-time']
    paginate_by = 10


# used to generate a bill object and save items
class SaleCreateView(View):                                                      
    template_name = 'sales/new_sale.html'

    def get(self, request):
        form = SaleForm(request.GET or None)
        formset = SaleItemFormset(request.GET or None)                          # renders an empty formset
        stocks = Stock.objects.filter(is_deleted=False)
        context = {
            'form'      : form,
            'formset'   : formset,
            'stocks'    : stocks
        }
        return render(request, self.template_name, context)

    def post(self, request):
        form = SaleForm(request.POST)
        formset = SaleItemFormset(request.POST)                                 # recieves a post method for the formset
        if form.is_valid() and formset.is_valid():
            # saves bill
            billobj = form.save(commit=False)
            billobj.save()     
            # create bill details object
            billdetailsobj = SaleBillDetails(billno=billobj)
            billdetailsobj.save()
            for form in formset:                                                # for loop to save each individual form as its own object
                # false saves the item and links bill to the item
                billitem = form.save(commit=False)
                billitem.billno = billobj                                       # links the bill object to the items
                # gets the stock item
                stock = get_object_or_404(Stock, name=billitem.stock.name)      
                # calculates the total price
                billitem.totalprice = billitem.perprice * billitem.quantity
                # updates quantity in stock db
                stock.quantity -= billitem.quantity   
                # saves bill item and stock
                stock.save()
                billitem.save()
            messages.success(request, "Sold items have been registered successfully")
            return redirect('sale-bill', billno=billobj.billno)
        form = SaleForm(request.GET or None)
        formset = SaleItemFormset(request.GET or None)
        context = {
            'form'      : form,
            'formset'   : formset,
        }
        return render(request, self.template_name, context)



# used to delete a bill object
class SaleDeleteView(SuccessMessageMixin, DeleteView):
    model = SaleBill
    template_name = "sales/delete_sale.html"
    success_url = '/transactions/sales'
    
    def delete(self, *args, **kwargs):
        self.object = self.get_object()
        items = SaleItem.objects.filter(billno=self.object.billno)
        for item in items:
            stock = get_object_or_404(Stock, name=item.stock.name)
            if stock.is_deleted == False:
                stock.quantity += item.quantity
                stock.save()
        messages.success(self.request, "Sale bill has been deleted successfully")
        return super(SaleDeleteView, self).delete(*args, **kwargs)




# used to display the purchase bill object
class PurchaseBillView(View):
    template_name = "bill/purchase_bill.html"
    bill_base = "bill/bill_base.html"

    def get(self, request, billno):
        bill = get_object_or_404(PurchaseBill, billno=billno)
        items = PurchaseItem.objects.filter(billno=billno).order_by('id')
        billdetails = get_object_or_404(PurchaseBillDetails, billno=billno)

        context = {
            'bill': bill,
            'items': items,
            'billdetails': billdetails,
            'bill_base': self.bill_base,
        }
        return render(request, self.template_name, context)

    def post(self, request, billno):
        form = PurchaseDetailsForm(request.POST)
        
   
        invoice = request.POST.get('invoice')
        if invoice:
            invoice = datetime.strptime(invoice, '%Y-%m-%d').date()


              
        
        if form.is_valid():
            
            billdetailsobj = get_object_or_404(PurchaseBillDetails, billno=billno)
            billdetailsobj.eway = form.cleaned_data['eway']
            billdetailsobj.veh = form.cleaned_data['veh']
           
            billdetailsobj.po = form.cleaned_data['po']
            billdetailsobj.receipt_date = form.cleaned_data['receipt_date']
           
            billdetailsobj.sup_invoice_no = form.cleaned_data['sup_invoice_no']
            if 'invoice' in form.cleaned_data:
                invoice = form.cleaned_data['invoice']
                if invoice:
                    billdetailsobj.invoice = invoice

            
            
            
            
            # Fetch total from PurchaseItem model
            purchase_items = PurchaseItem.objects.filter(billno=billno)
            total_price = purchase_items.aggregate(Sum('totalprice')).get('totalprice__sum', 0.0)
           
           

            # Apply calculations based on user input and fetched total
            if total_price:
                veh = form.cleaned_data['veh']
                total_price += veh

                cgst = total_price * 0.09
                sgst = total_price * 0.09
                igst = cgst + sgst
                grand_total = total_price + cgst + sgst

                # Assign calculated values to fields
                billdetailsobj.cgst = str(cgst)
                billdetailsobj.sgst = str(sgst)
                billdetailsobj.igst = str(igst)
                billdetailsobj.freight = str(veh)
                billdetailsobj.total = str(grand_total)
            
            billdetailsobj.invoice = invoice

            billdetailsobj.save()
            messages.success(request, "Bill details have been modified successfully")

        return redirect('purchase-bill', billno=billno)

# used to display the sale bill object
class SaleBillView(View):    
    model = SaleBill
    template_name = "bill/sale_bill.html"
    bill_base = "bill/bill_base.html"
    
    def get(self, request, billno):
        context = {
            'bill'          : SaleBill.objects.get(billno=billno),
            'items'         : SaleItem.objects.filter(billno=billno),
            'billdetails'   : SaleBillDetails.objects.get(billno=billno),
            'bill_base'     : self.bill_base,
        }
        return render(request, self.template_name, context)

    def post(self, request, billno):
        form = SaleDetailsForm(request.POST)
        if form.is_valid():
            billdetailsobj = SaleBillDetails.objects.get(billno=billno)
            
            billdetailsobj.eway = request.POST.get("eway")    
            billdetailsobj.veh = request.POST.get("veh")
            billdetailsobj.destination = request.POST.get("destination")
            billdetailsobj.po = request.POST.get("po")
            billdetailsobj.cgst = request.POST.get("cgst")
            billdetailsobj.sgst = request.POST.get("sgst")
            billdetailsobj.igst = request.POST.get("igst")
            billdetailsobj.cess = request.POST.get("cess")
            billdetailsobj.tcs = request.POST.get("tcs")
            billdetailsobj.total = request.POST.get("total")

            billdetailsobj.save()
            messages.success(request, "Bill details have been modified successfully")
        context = {
            'bill'          : SaleBill.objects.get(billno=billno),
            'items'         : SaleItem.objects.filter(billno=billno),
            'billdetails'   : SaleBillDetails.objects.get(billno=billno),
            'bill_base'     : self.bill_base,
        }
        return render(request, self.template_name, context)
 

def bill_list(request):
    search_query = request.GET.get('search', '')

    bills = PurchaseBillDetails.objects.all()

    if search_query:
        bills = bills.filter(Q(billno__billno__icontains=search_query))
        
    bills = bills.order_by('-billno__billno')

    # Rest of the pagination logic remains the same
    items_per_page = 10

    paginator = Paginator(bills, items_per_page)
    page = request.GET.get('page')

    try:
        bills_paginated = paginator.page(page)
    except PageNotAnInteger:
        bills_paginated = paginator.page(1)
    except EmptyPage:
        bills_paginated = paginator.page(paginator.num_pages)

    context = {
        'bills': bills_paginated,
        'search_query': search_query,  # Pass the search query back to the template
    }
    return render(request, 'purchases/bill_list.html', context)


def create_bom(request):
    if request.method == 'POST':
        bom_name = request.POST['name']
        code = request.POST['code']
        raw_materials = request.POST.getlist('raw_materials[]')
        quantities = request.POST.getlist('quantities[]')

        bom = BOM(name=bom_name,code=code)
        bom.save()

        for material, quantity in zip(raw_materials, quantities):
            bom_material = BOMRawMaterial(bom=bom,raw_material_id=material, quantity=quantity)
            bom_material.save()

        messages.success(request, 'BOM saved successfully.')  # Add success message

        return redirect('bom')  # Replace 'bom-list' with the appropriate URL name for the BOM list view

    raw_materials = Stock.objects.all()
    context = {'raw_materials': raw_materials}
    return render(request, 'production/create_bom.html', context)



def produce_item(request):
    if request.method == 'POST':
        bom_id = request.POST['bom']
        quantity = float(request.POST.get('quantities', 0)) 
        production_date = request.POST.get('production_date')# Convert quantity to float, defaulting to 0 if missing
        
        
        production_date = datetime.strptime(production_date, '%Y-%m-%d').date()
        bom = BOM.objects.get(id=bom_id)
        
        # Check if inventory is sufficient
        raw_materials = bom.raw_materials.all()
        insufficient_inventory = False
        insufficient_raw_materials = []
        for raw_material in raw_materials:
            bom_raw_material = BOMRawMaterial.objects.get(bom=bom, raw_material=raw_material)
            quantity_required = float(bom_raw_material.quantity) * quantity
            
            if raw_material.quantity < quantity_required:
                insufficient_raw_materials.append((raw_material.name, quantity_required))
                insufficient_inventory = True
        
        if insufficient_inventory:
            messages.error(request, 'Insufficient inventory for the following raw materials:')
            for material, required_quantity in insufficient_raw_materials:
                messages.error(request, f'- {material}: {required_quantity}')
            return redirect('production')  # Redirect to production list or desired URL
        
        # Deduct raw materials from inventory
        for raw_material in raw_materials:
            bom_raw_material = BOMRawMaterial.objects.get(bom=bom, raw_material=raw_material)
            quantity_required = float(bom_raw_material.quantity) * quantity
            raw_material.quantity -= quantity_required
            raw_material.save()
        
        # Save production data
        production = Production(bom=bom, quantity=quantity, production_date=production_date)
        # Generate the code_sfg
        production.save()
        
        # Update the total_qty field based on BOM name
        productions_with_same_bom = Production.objects.filter(bom__name=bom.name)
        total_quantity = productions_with_same_bom.aggregate(Sum('quantity'))['quantity__sum']
        production.total_qty = total_quantity
        production.save()
        
        messages.success(request, 'Order has been placed successfully')
        
        # Generate PDF
        pdf_data = generate_pdf(bom, quantity)
        
        # Return the PDF file as a response
        response = FileResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="production_report.pdf"'
        return response
    
    else:
        boms = BOM.objects.all()
        bom_quantities = {}
        productions = Production.objects.values('bom').annotate(total_quantity=Sum('quantity'))

        for production in productions:
            bom_id = production['bom']
            total_quantity = production['total_quantity']
            bom = BOM.objects.get(id=bom_id)
            
            bom_quantities[bom] = total_quantity
        
        return render(request, 'production/produce_item.html', {'boms': boms, 'bom_quantities': bom_quantities})
    
   


def generate_pdf(bom, quantity):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()

    elements = []

    # Production Details
    product_code = BOM.objects.latest("id").code
    production_id = Production.objects.latest('id').code_sfg
    product_name = Production.objects.latest('id').bom
    product_quantity = Production.objects.latest('id').quantity
    production_date = Production.objects.latest('id').production_date.strftime('%Y-%m-%d')

    elements.append(Paragraph("Instantina Flavours India Private Limited", styles['Heading1']))
    elements.append(Paragraph("Production Report SFG", styles['Heading1']))
    elements.append(Paragraph(f"Product ID: {product_code}", styles['Heading3']))
    elements.append(Paragraph(f"Batch No: {production_id}", styles['Heading3']))
    elements.append(Paragraph(f"Product Name: {product_name}", styles['Heading3']))
    elements.append(Paragraph(f"Production Quantity Kg: {product_quantity}", styles['Heading3']))
    elements.append(Paragraph(f"Date: {production_date}", styles['Normal']))
    elements.append(Spacer(1, 0.5 * inch))

    # Table Data
    data = [["Raw Material Code", "Raw Material", "Quantity Required (Kg)","RM Batch No","Best Before"]]
    total_quantity = 0  # Variable to hold the total quantity

    bom_raw_materials = BOMRawMaterial.objects.filter(bom=bom)
    for bom_raw_material in bom_raw_materials:
        quantity_required = round(bom_raw_material.quantity * quantity, 5)
        stock = bom_raw_material.raw_material
        data.append([stock.item_code, stock.name, str(quantity_required)])
        total_quantity += quantity_required  # Accumulate the total quantity

    # Add the row for total quantity
    data.append(["Total", "", str(total_quantity), "", ""])

    # Table Style (unchanged)
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
    ])
   

    table = Table(data, colWidths=[150, 150, 150, 100, 100])
    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)

    buffer.seek(0)
    return buffer



def bom_list(request):
    query = request.GET.get('query')
    boms = BOM.objects.all()
    
    if query:
        boms = boms.filter(name__icontains=query)
    
    context = {'boms': boms, 'query': query}
    return render(request, 'production/bom_list.html', context)




def update_bom(request, bom_id):
    try:
        bom = BOM.objects.get(id=bom_id)
    except BOM.DoesNotExist:
        messages.error(request, 'BOM not found.')
        return redirect('bom-list')

    if request.method == 'POST':
        bom_name = request.POST['name']
        code = request.POST['code']
        raw_materials = request.POST.getlist('raw_materials[]')
        quantities = request.POST.getlist('quantities[]')

        existing_raw_materials = BOMRawMaterial.objects.filter(bom=bom)

        # Update BOM name and code
        bom.name = bom_name
        bom.code = code
        bom.save()

        for material_id, quantity in zip(raw_materials, quantities):
            raw_material = Stock.objects.get(id=material_id)

            # Update existing raw material quantity or create new if not exist
            bom_raw_material, created = BOMRawMaterial.objects.get_or_create(
                bom=bom, raw_material=raw_material, defaults={'quantity': quantity}
            )
            if not created:
                bom_raw_material.quantity = quantity
                bom_raw_material.save()

        # Remove raw materials that were not included in the update
        existing_raw_materials.exclude(raw_material_id__in=raw_materials).delete()

        messages.success(request, 'BOM updated successfully.')
        return redirect('bom-list')

    raw_materials = Stock.objects.all()
    bom_raw_materials = BOMRawMaterial.objects.filter(bom=bom)
    context = {'bom': bom, 'bom_raw_materials': bom_raw_materials, 'raw_materials': raw_materials, 'updating': True}
    return render(request, 'production/update_bom.html', context)





def bom_listFG(request):
    query = request.GET.get('query')
    bom_raw_materials = FGSFG.objects.all()
    
    if query:
        bom_raw_materials = bom_raw_materials.filter(bom__name__icontains=query)
    
    paginator = Paginator(bom_raw_materials, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {'bom_raw_materials': page_obj, 'query': query}
    return render(request, 'production/bom_listfg.html', context)  


def produce_list(request):
    query = request.GET.get('query')
    productions = Production.objects.all()

    if query:
        productions = productions.filter(bom__name__icontains=query)

    # Calculate the sum of quantity for each `bom__name`
    productions = productions.values('bom__name').annotate(sum_quantity=Sum('quantity'))

    context = {'productions': productions, 'query': query}
    return render(request, 'production/production_list.html', context)




def create_productionfg(request):
    if request.method == 'POST':
        bom_id = request.POST['bom']
        sfg_name = request.POST['sfg']
        quantity_bom = float(request.POST.get('quantity_bom', 0))
        quantity_sfg = float(request.POST.get('quantity_sfg', 0))

        bom = BOM.objects.get(id=bom_id)
        sfg = Production.objects.filter(bom__name=sfg_name).first()

        if sfg is None:
            messages.error(request, f'Production with name "{sfg_name}" does not exist')
            return redirect('bomfg')

        # Check if inventory is sufficient for BOM
        raw_materials = bom.raw_materials.all()
        insufficient_inventory = False
        insufficient_raw_materials = []
        for raw_material in raw_materials:
            bom_raw_material = BOMRawMaterial.objects.get(bom=bom, raw_material=raw_material)
            quantity_required = float(bom_raw_material.quantity) * quantity_bom

            if raw_material.quantity < quantity_required:
                insufficient_raw_materials.append((raw_material.name, quantity_required))
                insufficient_inventory = True

        if insufficient_inventory:
            messages.error(request, 'Insufficient inventory for the following raw materials:')
            for material, required_quantity in insufficient_raw_materials:
                messages.error(request, f'- {material}: {required_quantity}')
            return redirect('bomfg')

        # Deduct raw materials from inventory for BOM
        for raw_material in raw_materials:
            bom_raw_material = BOMRawMaterial.objects.get(bom=bom, raw_material=raw_material)
            quantity_required = float(bom_raw_material.quantity) * quantity_bom
            raw_material.quantity -= quantity_required
            raw_material.save()

        # Check if SFG quantity is available
        if sfg.total_qty < quantity_sfg:
            messages.error(request, 'Insufficient SFG quantity')
            return redirect('bomfg')

        # Deduct quantity from SFG
        sfg.total_qty -= quantity_sfg
        sfg.save()
        
        # Update the total quantity of SFG
       

        # Save production data
        production = ProductionFG(bom=bom, quantity_bom=quantity_bom, sfg=sfg, quantity_sfg=quantity_sfg)
        production.save()

        messages.success(request, 'Order has been placed successfully')
        

        # Generate PDF
        pdf_data = generate_pdffg(bom, quantity_bom, quantity_sfg)

        # Return the PDF file as a response
        response = FileResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="production_report.pdf"'
        return response

    else:
        boms = BOM.objects.all()
        sfgs = Production.objects.values('bom__name').distinct()

        return render(request, 'production/create_bomFG.html', {'boms': boms, 'sfgs': sfgs})
    
    


def generate_pdffg(bom, quantity_bom, quantity_sfg):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()

    elements = []

    # Production Details
    productionfg = ProductionFG.objects.filter(bom=bom).latest('id')
    production_id = productionfg.bom.code
    product_name = productionfg.bom
    product_quantity = productionfg.quantity_bom
    product_code = productionfg.code_fg
    production_date = productionfg.production_date.strftime('%Y-%m-%d')

    elements.append(Paragraph("Instantina Flavours India Private Limited", styles['Heading1']))
    elements.append(Paragraph("Production Report FG", styles['Heading1']))
    elements.append(Paragraph(f"Product ID: {production_id}", styles['Heading3']))
    elements.append(Paragraph(f"Batch No.: {product_code}", styles['Heading3']))
    elements.append(Paragraph(f"Product Name: {product_name}", styles['Heading3']))
   
    elements.append(Paragraph(f"Production Quantity Kg: {product_quantity}" , styles['Heading3']))
    elements.append(Paragraph(f"Date: {production_date}", styles['Normal']))
    elements.append(Spacer(1, 0.8 * inch))

    # Table Data
    data = [["RM Code", "Raw Material", "Qty (Kg)", "SFG Name", "SFG Quantity (Kg)", "RM Batch no", "SFG Id", "Completion date"]]

    raw_materials = bom.raw_materials.all()
    sfg_name_added = False  # Flag to track if sfg_name is added to data

    for raw_material in raw_materials:
        bom_raw_material = BOMRawMaterial.objects.get(bom=bom, raw_material=raw_material)
        quantity_required = round(float(bom_raw_material.quantity) * quantity_bom, 5)

        sfg_name = productionfg.sfg.bom if productionfg.sfg and not sfg_name_added else ""
        sfg_quantity = productionfg.quantity_sfg if productionfg.sfg and not sfg_name_added else 0
        stock = bom_raw_material.raw_material
        data.append([stock.item_code, raw_material.name, str(quantity_required), sfg_name, str(sfg_quantity)])

        # Set sfg_name_added to True if sfg_name is added to data
        if sfg_name:
            sfg_name_added = True

    # Table Style
    table_style = TableStyle([
        
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
        
        
        # Table style definitions
    ])

    table = Table(data, colWidths=[80, 150, 80, 120, 80, 80, 80, 100])
    table.setStyle(table_style)
    elements.append(table)

    doc.build(elements)

    buffer.seek(0)
    return buffer



def produce_listfg(request):
    query = request.GET.get('query')
    leadtimes = Leadtimefg.objects.all()

    if query:
        leadtimes = leadtimes.filter(sfg__bom__name__icontains=query)

    paginator = Paginator(leadtimes, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {'leadtimes': page_obj, 'query': query}
    return render(request, 'production/produce_listfg.html', context)



def leadtimesfg(request):
    if request.method == 'POST':
        sfg_id = request.POST.get('sfg')
        rawcode = request.POST.get('rawcode')

        # Check if the SFG already has a Rawcode associated with it
        existing_leadtime = Leadtimesfg.objects.filter(sfg_id=sfg_id, Rawcode__isnull=False).first()
        if existing_leadtime:
            error_message = f"Rawcode already exists for SFG ID: {sfg_id}"
            # You can choose to display this error message in the template
            return render(request, 'production/leadtimesfg.html', {'error_message': error_message})

        # Create a new Leadtimesfg object
        new_leadtime = Leadtimesfg(sfg_id=sfg_id, Rawcode=rawcode)
        new_leadtime.save()

        # Redirect to a success page or display a success message
        return redirect('production-record')

    # If the request method is GET, retrieve the available SFGs
    sfg_list = Production.objects.filter(leadtimesfg__isnull=True)
    return render(request, 'production/leadtimesfg.html', {'sfg_list': sfg_list})



def leadtimelist(request):
    query = request.GET.get('query')
    leadtimes = Production.objects.order_by('-id')

    if query:
        leadtimes = leadtimes.filter(bom__name__icontains=query)

    paginator = Paginator(leadtimes, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {'leadtimes': page_obj, 'query': query}
    return render(request, 'production/leadtimelist.html', context)



def leadtimefg(request):
    if request.method == 'POST':
        sfg_id = request.POST.get('sfg')
        rawcode = request.POST.get('rawcode')

        # Check if the SFG already has a Rawcode associated with it
        existing_leadtime = Leadtimefg.objects.filter(sfg_id=sfg_id, Rawcode__isnull=False).first()
        if existing_leadtime:
            error_message = f"Rawcode already exists for SFG ID: {sfg_id}"
            # You can choose to display this error message in the template
            return render(request, 'production/leadtimesfg.html', {'error_message': error_message})

        # Create a new Leadtimesfg object
        new_leadtime = Leadtimefg(sfg_id=sfg_id, Rawcode=rawcode)
        new_leadtime.save()

        # Redirect to a success page or display a success message
        return redirect('production-record')

    # If the request method is GET, retrieve the available SFGs
    sfg_list = ProductionFG.objects.filter(leadtimefg__isnull=True)
    return render(request, 'production/leadtimefg.html', {'sfg_list': sfg_list})



def leadtimelistfg(request):
    query = request.GET.get('query')
    leadtimes = fgproduction.objects.order_by('-id')

    if query:
        leadtimes = leadtimes.filter(bom__name__icontains=query)

    paginator = Paginator(leadtimes, 10)  # Show 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {'leadtimes': page_obj, 'query': query}
    return render(request, 'production/leadtimelistfg.html', context)



def bom_details(request, bom_id):
    bom = get_object_or_404(BOM, pk=bom_id)
    bom_raw_materials = bom.bomrawmaterial_set.all()
    context = {
        'bom': bom,
        'bom_raw_materials': bom_raw_materials,
        'row_count': len(bom_raw_materials),
    }
    return render(request, 'production/bom_details.html', context)





def create_sfg_bom(request):
    if request.method == 'POST':
       
        sfg_id = request.POST['sfg']
        quantity_sfg = request.POST['quantity_sfg']

     
        sfg = Production.objects.get(id=sfg_id)

        sfg_bom = SfgBom( sfg=sfg, quantity_sfg=quantity_sfg)
        sfg_bom.save()

        messages.success(request, 'BOM saved successfully.')  # Add success message

        return redirect('sfgbom')  # Replace 'bom_list' with the appropriate URL name for the BOM list view

    else:
        boms = BOM.objects.all()
        productions = Production.objects.all()

        context = {
            'boms': boms,
            'productions': productions,
        }

        return render(request, 'production/bom_sfg.html', context)





def create_sfgfinal(request):
    if request.method == 'POST':
        # Get the selected BOM and SfgBom names from the form data
        bom_name = request.POST.get('bom')
        sfg_name = request.POST.get('sfg')

        # Retrieve the BOM and SfgBom objects based on the selected names
        bom = BOM.objects.get(name=bom_name)
        sfg = SfgBom.objects.filter(name=sfg_name).first()

        if sfg:
            # Create and save the Sfgfinal object
            sfgfinal = Sfgfinal(bom=bom, sfg=sfg)
            sfgfinal.save()

            # Redirect to a success page or do something else
            return redirect('success')

    # Retrieve the available BOM and unique SfgBom names for display in the form
    boms = BOM.objects.all()
    sfg_names = SfgBom.objects.values('sfg').distinct()

    context = {
        'boms': boms,
        'sfg_names': sfg_names,
    }
    return render(request, 'production/create_sfgfinal.html', context)



# ... (existing code)

def create_fgsfgbom(request):
    if request.method == 'POST':
        name = request.POST['name']
        code = request.POST['code']
        sfg_ids = request.POST.getlist('sfg[]')
        quantities_sfg = request.POST.getlist('quantity_sfg[]')
        raw_materials = request.POST.getlist('raw_materials[]')
        quantities_raw = request.POST.getlist('quantities[]')

        # Save Finished Goods (FG) entry
        fgsfgnew = FGSFGNEW(name=name, code=code)
        fgsfgnew.save()

        # Save Semi-Finished Goods (SFG) for the FG entry
        for sfg_id, quantity_sfg in zip(sfg_ids, quantities_sfg):
            sfg = BOM.objects.get(id=sfg_id)
            multiple_sfg = MultipleSFG(fgsfgnew=fgsfgnew, sfg=sfg, quantity_sfg=quantity_sfg)
            multiple_sfg.save()

        # Save Raw Materials for the FG entry
        for material, quantity_raw in zip(raw_materials, quantities_raw):
            raw_material = Stock.objects.get(id=material)
            raw_material_entry = RawMaterialEntry(fgsfgnew=fgsfgnew, raw_material=raw_material, quantity_raw=quantity_raw)
            raw_material_entry.save()

        messages.success(request, 'FG BOM SAVED.')
        return redirect('production-fgsfgbomlist')

    sfg_list = BOM.objects.all()
    raw_materials = Stock.objects.all()
    context = {
        'sfg_list': sfg_list,
        'raw_materials': raw_materials,
    }

    return render(request, 'production/create_fgsfgbom.html', context)



def update_fgsfgbom(request, fgsfg_id):
    try:
        fgsfg = FGSFGNEW.objects.get(id=fgsfg_id)
    except FGSFGNEW.DoesNotExist:
        messages.error(request, 'FG BOM not found.')
        return redirect('production-fgsfgbomlist')
        

    if request.method == 'POST':        
        name=request.POST['name']
        code=request.POST['code']
        sfg_ids = request.POST.getlist('sfg[]')
        quantities_sfg = request.POST.getlist('quantity_sfg[]')
        raw_materials = request.POST.getlist('raw_materials[]')
        quantities_raw = request.POST.getlist('quantities[]')
        
        
        fgsfg.name = name
        fgsfg.code = code
        fgsfg.save()

        # Clear existing SFG and Raw Material entries associated with the FG
        MultipleSFG.objects.filter(fgsfgnew=fgsfg).delete()
        RawMaterialEntry.objects.filter(fgsfgnew=fgsfg).delete()

        # Save Semi-Finished Goods (SFG) for the FG entry
        for sfg_id, quantity_sfg in zip(sfg_ids, quantities_sfg):
            sfg = BOM.objects.get(id=sfg_id)
            multiple_sfg = MultipleSFG(fgsfgnew=fgsfg, sfg=sfg, quantity_sfg=quantity_sfg)
            multiple_sfg.save()

        # Save Raw Materials for the FG entry
        for material, quantity_raw in zip(raw_materials, quantities_raw):
            raw_material = Stock.objects.get(id=material)
            raw_material_entry = RawMaterialEntry(fgsfgnew=fgsfg, raw_material=raw_material, quantity_raw=quantity_raw)
            raw_material_entry.save()

        messages.success(request, 'FG BOM updated successfully.')
        return redirect('production-fgsfgbomlist')

    sfg_list = BOM.objects.all()
    raw_materials = Stock.objects.all()
    existing_sfg_entries = MultipleSFG.objects.filter(fgsfgnew=fgsfg)
    existing_raw_material_entries = RawMaterialEntry.objects.filter(fgsfgnew=fgsfg)
    context = {
        'fgsfg': fgsfg,
        'sfg_list': sfg_list,
        'raw_materials': raw_materials,
        'existing_sfg_entries': existing_sfg_entries,
        'existing_raw_material_entries': existing_raw_material_entries,
        'updating': True,
    }

    return render(request, 'production/update_fgsfgbom.html', context)




def sfg_production_view(request):
    if request.method == 'POST':
        bom_id = request.POST['bom']
        quantity = float(request.POST.get('quantity', 0))
        production_date = request.POST.get('production_date')
        
        production_date = datetime.strptime(production_date, '%Y-%m-%d').date()

        bom = FGSFGNEW.objects.get(id=bom_id)
        raw_material_entries = bom.rawmaterialentry_set.all()
        sfg_entries = MultipleSFG.objects.filter(fgsfgnew=bom)
        sfg_quantity_needed = {}

        # Calculate total required quantity for each SFG
        for entry in sfg_entries:
            sfg_name = entry.sfg.name
            sfg_quantity_needed[sfg_name] = sfg_quantity_needed.get(sfg_name, 0) + (entry.quantity_sfg * quantity)

        # Check if inventory is sufficient for all raw materials
        raw_materials = bom.raw_materials.all()
        insufficient_inventory = False
        insufficient_raw_materials = []
        for raw_material in raw_materials:
            bom_raw_material = RawMaterialEntry.objects.get(fgsfgnew=bom, raw_material=raw_material)

            quantity_required = float(bom_raw_material.quantity_raw) * quantity

            if raw_material.quantity < quantity_required:
                insufficient_raw_materials.append((raw_material.name, quantity_required))
                insufficient_inventory = True

        if insufficient_inventory:
            messages.error(request, 'Insufficient inventory for the following raw materials:')
            for material, required_quantity in insufficient_raw_materials:
                messages.error(request, f'- {material}: {required_quantity}')
            return redirect('production')

        # Check if sufficient SFG quantity is available for all SFGs
        for sfg_name, sfg_quantity in sfg_quantity_needed.items():
            # Retrieve suitable SFG productions in ascending order of quantity
            sfgs = Production.objects.filter(bom__name=sfg_name).order_by('quantity')

            # Check if any suitable SFG production is available
            for sfg in sfgs:
                if sfg.quantity >= sfg_quantity:
                    # Sufficient SFG quantity is available, deduct the required quantity from the production
                    sfg.quantity -= sfg_quantity
                    sfg.save()

                    # Create and save the FGProduction object
                    production = fgproduction(bom=bom, quantity=quantity,production_date=production_date)
                    production.save()

                    # Generate PDF
                    pdf_data = generate_pdfsfgfg(bom, quantity)
                    
                    messages.success(request, 'Production order placed')

                    # Return the PDF file as a response
                    response = FileResponse(pdf_data, content_type='application/pdf')
                    response['Content-Disposition'] = 'attachment; filename="production_report.pdf"'
                    return response
                else:
                    # Insufficient SFG quantity in this production, try the next production
                    sfg_quantity -= sfg.quantity
                    sfg.quantity = 0
                    sfg.save()

        # If this point is reached, it means there is no sufficient SFG quantity in any production for at least one SFG
        messages.error(request, 'Insufficient SFG quantity')
        return redirect('production-sfg')

    else:
        boms = FGSFGNEW.objects.all()
        return render(request, 'production/production_sfg.html', {'boms': boms})




def generate_pdfsfgfg(bom, quantity):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()

    elements = []

    # Production Details
    fg_prod = fgproduction.objects.filter(bom=bom).latest('id')
    production_id = fg_prod.bom.code
    product_name = fg_prod.bom.name
    product_quantity = fg_prod.quantity
    product_code = fg_prod.code_fg
    production_date = fg_prod.production_date.strftime('%Y-%m-%d')

    elements.append(Paragraph("Instantina Flavours India Private Limited", styles['Heading1']))
    elements.append(Paragraph("Production Report FG", styles['Heading1']))
    elements.append(Paragraph(f"Product ID: {production_id}", styles['Heading3']))
    elements.append(Paragraph(f"Batch No.: {product_code}", styles['Heading3']))
    elements.append(Paragraph(f"Product Name: {product_name}", styles['Heading3']))

    elements.append(Paragraph(f"Production Quantity Kg: {product_quantity}", styles['Heading3']))
    elements.append(Paragraph(f"Date: {production_date}", styles['Normal']))
    elements.append(Spacer(1, 0.8 * inch))

    # Table Data for Raw Materials
    data_raw_material = [["RM Code", "Raw Material", "Qty (Kg)"]]
    raw_material_entries = bom.rawmaterialentry_set.all()
    for entry in raw_material_entries:
        raw_material = entry.raw_material
        quantity_raw = entry.quantity_raw
        quantity_required = quantity_raw * quantity
        data_raw_material.append([raw_material.item_code, raw_material.name, str(quantity_required)])

    # Table Style for Raw Materials
    table_style_raw_material = TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
    ])

    table_raw_material = Table(data_raw_material, colWidths=[200,200,200])
    table_raw_material.setStyle(table_style_raw_material)
    elements.append(Paragraph("Raw Materials", styles['Heading2']))
    elements.append(table_raw_material)

    # Table Data for SFGs
    data_sfg = [["SFG Name", "SFG Quantity (Kg)"]]
    sfg_entries = MultipleSFG.objects.filter(fgsfgnew=bom)
    for sfg_entry in sfg_entries:
        sfg_name = sfg_entry.sfg.name
        sfg_quantity = sfg_entry.quantity_sfg * quantity
        data_sfg.append([sfg_name, str(sfg_quantity)])

    # Table Style for SFGs
    table_style_sfg = TableStyle([
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
    ])

    table_sfg = Table(data_sfg, colWidths=[200,200])
    table_sfg.setStyle(table_style_sfg)
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph("Semi-Finished Goods (SFGs)", styles['Heading2']))
    elements.append(table_sfg)

    doc.build(elements)

    buffer.seek(0)
    return buffer



       
def fgsfgbom_list(request):
    query = request.GET.get('query')
    boms = FGSFGNEW.objects.all()
    
    if query:
        boms = boms.filter(name__icontains=query)
    
    context = {'boms': boms, 'query': query}
    return render(request, 'production/fgsfgbom_list.html', context)




def bom_detailsfgsfg(request, bom_id):
    fgsfgnew = get_object_or_404(FGSFGNEW, pk=bom_id)
    raw_materials = fgsfgnew.raw_materials.through.objects.filter(fgsfgnew=fgsfgnew)
    sfgs = fgsfgnew.sfg.through.objects.filter(fgsfgnew=fgsfgnew)

    context = {
        'fgsfgnew': fgsfgnew,
        'raw_materials': raw_materials,
        'sfgs': sfgs,
    }

    return render(request, 'production/bom_detailsfgsfg.html', context)



